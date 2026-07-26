# routes/stats.py
from __future__ import annotations

from collections import Counter, defaultdict
from io import BytesIO
import json
import re
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.student_access import active_student_code_set_for_teacher, list_student_codes_for_teacher
from services.student_work_store import list_teacher_student_work_metadata
from services.teacher_profiles import get_teacher
from core.config import RUNS_ROOT
from core.ai_clients.ai_usage_logger import usage_log_dir
from core.security import require_teacher

router = APIRouter(prefix="/routes", tags=["stats"])

SUBMISSIONS_XLSX = RUNS_ROOT / "student_submissions.xlsx"
SHEET = "submissions"

UNMATCHED_EXAM_LABEL = "Unmatched / not graded yet"
ALL_VOUCHERS = "__all__"


def _read_rows() -> list[dict[str, Any]]:
    if not SUBMISSIONS_XLSX.exists():
        return []

    wb = load_workbook(SUBMISSIONS_XLSX)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    out = []
    for r in rows[1:]:
        d = {}
        for i, key in enumerate(header):
            if not key:
                continue
            d[key] = r[i] if i < len(r) else None
        out.append(d)
    return out


def _teacher_id_from_session(session: dict[str, Any]) -> str:
    teacher_id = (session.get("teacher_id") or session.get("sub") or "").strip()
    if not teacher_id or teacher_id == "teacher":
        raise HTTPException(status_code=400, detail="Legacy teacher sessions cannot use dashboard upload analytics.")
    return teacher_id


def _default_voucher_for_teacher(teacher_id: str) -> str:
    teacher = get_teacher(teacher_id) or {}
    return str(
        teacher.get("voucher_id")
        or teacher.get("voucher_hash")
        or teacher.get("voucher_code")
        or "voucher_default"
    ).strip() or "voucher_default"


def _roster_for_teacher(teacher_id: str) -> list[dict[str, Any]]:
    default_voucher = _default_voucher_for_teacher(teacher_id)
    roster: list[dict[str, Any]] = []

    for rec in list_student_codes_for_teacher(teacher_id):
        if rec.get("status") != "active":
            continue
        code = str(rec.get("code") or "").strip()
        if not code:
            continue

        voucher_id = str(
            rec.get("voucher_id")
            or rec.get("voucher_hash")
            or rec.get("voucher_code")
            or default_voucher
        ).strip() or default_voucher

        roster.append(
            {
                "code": code,
                "voucher_id": voucher_id,
                "course_label": str(rec.get("course_label") or "").strip(),
                "student_name": str(rec.get("student_name") or "").strip(),
                "student_email": str(rec.get("student_email") or "").strip(),
                "note": str(rec.get("note") or "").strip(),
                "last_used_at": str(rec.get("last_used_at") or ""),
                "uses": int(rec.get("uses") or 0),
            }
        )

    return roster


def _exam_id_for_work(meta: dict[str, Any]) -> tuple[str, bool]:
    exam_id = str(meta.get("exam_id") or "").strip()
    if exam_id:
        return exam_id, True
    return UNMATCHED_EXAM_LABEL, False


def _best_upload_time(meta: dict[str, Any]) -> str:
    return str(meta.get("saved_at") or meta.get("metadata_saved_at") or "").strip()


def _best_last_time(meta: dict[str, Any]) -> str:
    feedback = meta.get("feedback") if isinstance(meta.get("feedback"), dict) else {}
    return str(
        meta.get("graded_at")
        or feedback.get("saved_at")
        or meta.get("updated_at")
        or meta.get("saved_at")
        or ""
    ).strip()


def _metadata_file_by_kind(
    meta: dict[str, Any],
    *kinds: str,
) -> str:
    wanted = set(kinds)

    for item in meta.get("files") or []:
        if not isinstance(item, dict):
            continue

        if str(item.get("kind") or "") not in wanted:
            continue

        filename = str(item.get("filename") or "").strip()

        if filename:
            return filename

    return ""


def _review_files_for_meta(
    meta: dict[str, Any],
) -> dict[str, str]:
    feedback = (
        meta.get("feedback")
        if isinstance(meta.get("feedback"), dict)
        else {}
    )

    original_filename = _metadata_file_by_kind(
        meta,
        "original",
        "student_tex",
    )

    ocr_tex_filename = _metadata_file_by_kind(
        meta,
        "ocr_tex",
    )

    if not ocr_tex_filename:
        primary_filename = str(
            meta.get("primary_filename")
            or ""
        ).strip()

        if primary_filename.lower().endswith((".tex", ".txt")):
            ocr_tex_filename = primary_filename

    graded_result_filename = _metadata_file_by_kind(
        meta,
        "graded_result_json",
    )

    if not graded_result_filename:
        graded_result_filename = str(
            feedback.get("structured_json_filename")
            or ""
        ).strip()

    if not graded_result_filename:
        graded_result_filename = _metadata_file_by_kind(
            meta,
            "graded_result_tex",
            "graded_feedback",
        )

    if not graded_result_filename:
        graded_result_filename = str(
            feedback.get("filename")
            or ""
        ).strip()

    return {
        "original_filename": original_filename,
        "ocr_tex_filename": ocr_tex_filename,
        "graded_result_filename": graded_result_filename,
    }


def _safe_filename_part(value: str, fallback: str = "upload_log") -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", (value or "").strip()).strip("._-")
    return cleaned[:80] or fallback


def _read_teacher_ai_usage(
    teacher_id: str,
    teacher_codes: set[str],
) -> tuple[Counter, Counter]:
    """
    Aggregate GPT and Gemini token usage from the shared JSONL usage logs.

    New records are selected using teacher_id. Records created before teacher
    attribution was added can still be used when their student_code belongs
    to this teacher.
    """
    gemini_by_code: Counter = Counter()
    gpt_by_code: Counter = Counter()

    gemini_providers = {
        "google",
        "gemini",
        "google_ai_studio",
        "aistudio",
    }
    gpt_providers = {
        "openai",
        "gpt",
        "chatgpt",
        "openai_gpt",
        "azure_openai",
    }

    try:
        log_root = usage_log_dir()
    except Exception:
        return gemini_by_code, gpt_by_code

    for log_path in sorted(log_root.glob("ai_usage_*.jsonl")):
        try:
            lines = log_path.read_text(
                encoding="utf-8",
                errors="replace",
            ).splitlines()
        except Exception:
            continue

        for line in lines:
            if not line.strip():
                continue

            try:
                record = json.loads(line)
            except Exception:
                continue

            record_teacher_id = str(
                record.get("teacher_id")
                or ""
            ).strip()

            student_code = str(
                record.get("student_code")
                or record.get("code")
                or ""
            ).strip()

            belongs_to_teacher = (
                record_teacher_id == teacher_id
                or (
                    not record_teacher_id
                    and student_code in teacher_codes
                )
            )

            if not belongs_to_teacher:
                continue

            # The current chart is per student code. Teacher test-mode calls
            # without a student code are intentionally excluded.
            if not student_code or student_code not in teacher_codes:
                continue

            try:
                total_tokens = int(
                    record.get("total_tokens")
                    or 0
                )
            except (TypeError, ValueError):
                total_tokens = 0

            if total_tokens <= 0:
                continue

            provider = str(
                record.get("provider")
                or ""
            ).strip().lower()

            if provider in gemini_providers:
                gemini_by_code[student_code] += total_tokens
            elif provider in gpt_providers:
                gpt_by_code[student_code] += total_tokens

    return gemini_by_code, gpt_by_code


def _build_teacher_upload_dashboard(teacher_id: str) -> dict[str, Any]:
    rows = _read_rows()
    teacher_codes = active_student_code_set_for_teacher(teacher_id)

    # Keep legacy usage/token charts working, but only for this teacher's codes.
    rows = [r for r in rows if (str(r.get("code") or "").strip() in teacher_codes)]

    roster = _roster_for_teacher(teacher_id)
    default_voucher = _default_voucher_for_teacher(teacher_id)

    roster_by_voucher: dict[str, list[dict[str, Any]]] = defaultdict(list)
    roster_by_code: dict[str, dict[str, Any]] = {}

    for rec in roster:
        roster_by_voucher[rec["voucher_id"]].append(rec)
        roster_by_code[rec["code"]] = rec

    metadata_items = list_teacher_student_work_metadata(teacher_id)

    per_day: dict[str, dict[str, Any]] = defaultdict(lambda: {"count": 0, "students": set()})
    recent_uploads: list[dict[str, Any]] = []
    work_groups: dict[tuple[str, str], dict[str, Any]] = {}

    for meta in metadata_items:
        code = str(meta.get("student_code") or "").strip()
        if not code:
            continue

        roster_rec = roster_by_code.get(code, {})
        voucher_id = str(meta.get("voucher_id") or roster_rec.get("voucher_id") or default_voucher).strip() or default_voucher
        exam_id, is_matched = _exam_id_for_work(meta)
        upload_at = _best_upload_time(meta)
        last_at = _best_last_time(meta)
        work_id = str(meta.get("work_id") or "").strip()
        review_files = _review_files_for_meta(meta)
        day = upload_at[:10] if upload_at else "unknown"

        per_day[day]["count"] += 1
        per_day[day]["students"].add(code)

        key = (voucher_id, exam_id)
        if key not in work_groups:
            work_groups[key] = {
                "voucher_id": voucher_id,
                "exam_id": exam_id,
                "is_matched_exam": is_matched,
                "upload_count": 0,
                "last_upload_at": "",
                "students": {},
            }

        group = work_groups[key]
        group["upload_count"] += 1

        if upload_at and upload_at > str(group.get("last_upload_at") or ""):
            group["last_upload_at"] = upload_at

        students = group["students"]
        if code not in students:
            students[code] = {
                "code": code,
                "course_label": str(roster_rec.get("course_label") or ""),
                "attempts": 0,
                "first_uploaded_at": upload_at,
                "last_uploaded_at": upload_at,
                "last_graded_at": "",
                "status": "",
                "source_filename": "",
                "latest_work_id": "",
                "original_filename": "",
                "ocr_tex_filename": "",
                "graded_result_filename": "",
                "work_ids": [],
            }

        student_row = students[code]
        student_row["attempts"] += 1

        if upload_at and (not student_row["first_uploaded_at"] or upload_at < student_row["first_uploaded_at"]):
            student_row["first_uploaded_at"] = upload_at

        if upload_at and upload_at > str(student_row.get("last_uploaded_at") or ""):
            student_row["last_uploaded_at"] = upload_at
            student_row["status"] = str(meta.get("status") or "")
            student_row["source_filename"] = str(meta.get("source_filename") or "")
            student_row["latest_work_id"] = work_id
            student_row.update(review_files)

        if work_id and not student_row.get("latest_work_id"):
            student_row["latest_work_id"] = work_id
            student_row.update(review_files)

        if last_at and last_at > str(student_row.get("last_graded_at") or ""):
            student_row["last_graded_at"] = last_at

        if work_id:
            student_row["work_ids"].append(work_id)

        recent_uploads.append(
            {
                "voucher_id": voucher_id,
                "course_label": str(roster_rec.get("course_label") or ""),
                "student_code": code,
                "exam_id": exam_id,
                "is_matched_exam": is_matched,
                "uploaded_at": upload_at,
                "last_graded_at": last_at,
                "status": str(meta.get("status") or ""),
                "kind": str(meta.get("kind") or ""),
                "ocr_provider": str(meta.get("ocr_provider") or ""),
                "grading_provider": str(meta.get("grading_provider") or ""),
                "source_filename": str(meta.get("source_filename") or ""),
                "work_id": str(meta.get("work_id") or ""),
            }
        )

    work_progress: list[dict[str, Any]] = []
    voucher_ids = set(roster_by_voucher.keys()) | {str(m.get("voucher_id") or default_voucher) for m in metadata_items}

    for (voucher_id, exam_id), group in work_groups.items():
        roster_for_voucher = roster_by_voucher.get(voucher_id, [])
        total_students = len(roster_for_voucher)

        uploaded_codes = set(group["students"].keys())
        if total_students <= 0:
            total_students = len(uploaded_codes)

        missing_codes = [r["code"] for r in roster_for_voucher if r["code"] not in uploaded_codes]
        uploaded_count = len(uploaded_codes)
        completion_pct = round((uploaded_count / total_students) * 100, 1) if total_students else 0.0

        work_progress.append(
            {
                "voucher_id": voucher_id,
                "exam_id": exam_id,
                "is_matched_exam": bool(group.get("is_matched_exam")),
                "uploaded_students": uploaded_count,
                "total_students": total_students,
                "completion_pct": completion_pct,
                "upload_count": int(group.get("upload_count") or 0),
                "last_upload_at": str(group.get("last_upload_at") or ""),
                "students_uploaded": sorted(
                    group["students"].values(),
                    key=lambda x: x.get("last_uploaded_at") or "",
                    reverse=True,
                ),
                "students_missing": missing_codes,
            }
        )

    work_progress.sort(key=lambda x: (str(x.get("voucher_id") or ""), str(x.get("exam_id") or "")))

    student_progress_by_code: dict[str, dict[str, Any]] = {}
    for rec in roster:
        student_progress_by_code[rec["code"]] = {
            "code": rec["code"],
            "voucher_id": rec.get("voucher_id", ""),
            "course_label": rec.get("course_label", ""),
            "student_name": rec.get("student_name", ""),
            "student_email": rec.get("student_email", ""),
            "note": rec.get("note", ""),
            "created_at": rec.get("created_at", ""),
            "last_used_at": rec.get("last_used_at", ""),
            "uses": int(rec.get("uses") or 0),
            "works": [],
            "total_attempts": 0,
            "uploaded_work_count": 0,
        }

    for work in work_progress:
        for uploaded in work.get("students_uploaded", []) or []:
            code = str(uploaded.get("code") or "").strip()
            if not code:
                continue

            row = student_progress_by_code.setdefault(
                code,
                {
                    "code": code,
                    "voucher_id": work.get("voucher_id", ""),
                    "course_label": str(uploaded.get("course_label") or ""),
                    "student_name": "",
                    "student_email": "",
                    "note": "",
                    "created_at": "",
                    "last_used_at": "",
                    "uses": 0,
                    "works": [],
                    "total_attempts": 0,
                    "uploaded_work_count": 0,
                },
            )

            attempts = int(uploaded.get("attempts") or 0)
            row["works"].append(
                {
                    "exam_id": work.get("exam_id", ""),
                    "voucher_id": work.get("voucher_id", ""),
                    "is_matched_exam": bool(work.get("is_matched_exam")),
                    "attempts": attempts,
                    "first_uploaded_at": uploaded.get("first_uploaded_at", ""),
                    "last_uploaded_at": uploaded.get("last_uploaded_at", ""),
                    "last_graded_at": uploaded.get("last_graded_at", ""),
                    "status": uploaded.get("status", ""),
                    "source_filename": uploaded.get("source_filename", ""),
                    "work_id": uploaded.get("latest_work_id", ""),
                    "original_filename": uploaded.get("original_filename", ""),
                    "ocr_tex_filename": uploaded.get("ocr_tex_filename", ""),
                    "graded_result_filename": uploaded.get(
                        "graded_result_filename",
                        "",
                    ),
                }
            )
            row["total_attempts"] = int(row.get("total_attempts") or 0) + attempts
            row["uploaded_work_count"] = int(row.get("uploaded_work_count") or 0) + 1

    student_progress = list(student_progress_by_code.values())
    for row in student_progress:
        row["works"].sort(key=lambda x: str(x.get("last_uploaded_at") or ""), reverse=True)
    student_progress.sort(key=lambda x: (str(x.get("course_label") or ""), str(x.get("code") or "")))

    vouchers = []
    for voucher_id in sorted(voucher_ids or {default_voucher}):
        roster_for_voucher = roster_by_voucher.get(voucher_id, [])
        labels = [
            str(r.get("course_label") or "").strip()
            for r in roster_for_voucher
            if str(r.get("course_label") or "").strip()
        ]

        upload_count = sum(1 for item in recent_uploads if item.get("voucher_id") == voucher_id)
        work_count = sum(
            1 for item in work_progress
            if item.get("voucher_id") == voucher_id and item.get("is_matched_exam")
        )
        last_uploads = [
            str(item.get("uploaded_at") or "")
            for item in recent_uploads
            if item.get("voucher_id") == voucher_id
        ]

        vouchers.append(
            {
                "voucher_id": voucher_id,
                "course_label": labels[0] if labels else "Course",
                "total_students": len(roster_for_voucher),
                "upload_count": upload_count,
                "work_count": work_count,
                "last_upload_at": max(last_uploads) if last_uploads else "",
            }
        )

    uploads_per_day = [
        {"day": day, "count": int(data["count"]), "unique_students": len(data["students"])}
        for day, data in sorted(per_day.items())
        if day and day != "unknown"
    ]

    if not uploads_per_day:
        legacy_per_day = Counter()
        for r in rows:
            ts = r.get("timestamp")
            if ts:
                legacy_per_day[str(ts)[:10]] += 1
        uploads_per_day = [
            {"day": d, "count": legacy_per_day[d], "unique_students": 0}
            for d in sorted(legacy_per_day.keys())
        ]

    codes_per_exam = [
        {
            "exam_id": x["exam_id"],
            "unique_codes": int(x["uploaded_students"]),
            "submissions": int(x["upload_count"]),
        }
        for x in work_progress
        if x.get("is_matched_exam")
    ]

    # Read token usage from the shared AI usage JSONL logs.
    # The helper separates GPT and Gemini usage and attributes each call
    # to the correct teacher and student code.
    gemini_by_code, gpt_by_code = _read_teacher_ai_usage(
        teacher_id=teacher_id,
        teacher_codes=teacher_codes,
    )

    all_token_codes = sorted(
        set(gemini_by_code.keys()) | set(gpt_by_code.keys())
    )

    gemini_tokens_per_code = [
        {
            "code": code,
            "tokens": int(gemini_by_code.get(code, 0)),
        }
        for code in all_token_codes
    ]

    gpt_tokens_per_code = [
        {
            "code": code,
            "tokens": int(gpt_by_code.get(code, 0)),
        }
        for code in all_token_codes
    ]

    recent_uploads.sort(key=lambda x: x.get("uploaded_at") or "", reverse=True)

    return {
        "ok": True,
        "roster": {
            "total_students": len(roster),
            "active_codes": [r["code"] for r in roster],
        },
        "vouchers": vouchers,
        "uploads_per_day": uploads_per_day,
        "usage_per_day": uploads_per_day,
        "work_progress": work_progress,
        "student_progress": student_progress,
        "codes_per_exam": codes_per_exam,
        "gemini_tokens_per_code": gemini_tokens_per_code,
        "gpt_tokens_per_code": gpt_tokens_per_code,
        "recent_uploads": recent_uploads,
    }


@router.get("/teacher_dashboard")
@router.get("/stats/teacher_dashboard")
def teacher_dashboard(_session: dict = Depends(require_teacher)):
    teacher_id = _teacher_id_from_session(_session)
    return _build_teacher_upload_dashboard(teacher_id)


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="F39A1E")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col = column_cells[0].column

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 42)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _append_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def _make_upload_log_workbook(payload: dict[str, Any], voucher_id: str) -> Workbook:
    selected_voucher = (voucher_id or ALL_VOUCHERS).strip()
    selected_all = selected_voucher in ("", ALL_VOUCHERS)

    progress = [
        x for x in payload.get("work_progress", [])
        if selected_all or x.get("voucher_id") == selected_voucher
    ]

    recent_uploads = [
        x for x in payload.get("recent_uploads", [])
        if selected_all or x.get("voucher_id") == selected_voucher
    ]

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    _append_rows(
        ws,
        [
            ["Voucher", "Work / exam", "Uploaded students", "Total active students", "Completion", "Upload attempts", "Last upload"],
            *[
                [
                    x.get("voucher_id", ""),
                    x.get("exam_id", ""),
                    x.get("uploaded_students", 0),
                    x.get("total_students", 0),
                    (float(x.get("completion_pct") or 0) / 100.0),
                    x.get("upload_count", 0),
                    x.get("last_upload_at", ""),
                ]
                for x in progress
            ],
        ],
    )

    for cell in ws["E"][1:]:
        cell.number_format = "0%"

    _style_header(ws)
    _autosize(ws)

    log_ws = wb.create_sheet("Upload log")
    attempt_counts = Counter(
        (x.get("voucher_id", ""), x.get("exam_id", ""), x.get("student_code", ""))
        for x in recent_uploads
    )

    _append_rows(
        log_ws,
        [
            [
                "Voucher",
                "Course / group",
                "Student code",
                "Work / exam",
                "Uploaded at",
                "Last graded / updated at",
                "Attempts for this student/work",
                "Status",
                "Upload type",
                "OCR provider",
                "Grading provider",
                "Source filename",
                "Work ID",
            ],
            *[
                [
                    x.get("voucher_id", ""),
                    x.get("course_label", ""),
                    x.get("student_code", ""),
                    x.get("exam_id", ""),
                    x.get("uploaded_at", ""),
                    x.get("last_graded_at", ""),
                    attempt_counts[(x.get("voucher_id", ""), x.get("exam_id", ""), x.get("student_code", ""))],
                    x.get("status", ""),
                    x.get("kind", ""),
                    x.get("ocr_provider", ""),
                    x.get("grading_provider", ""),
                    x.get("source_filename", ""),
                    x.get("work_id", ""),
                ]
                for x in recent_uploads
            ],
        ],
    )

    _style_header(log_ws)
    _autosize(log_ws)

    missing_ws = wb.create_sheet("Missing by work")
    missing_rows: list[list[Any]] = []

    for x in progress:
        for code in x.get("students_missing", []) or []:
            missing_rows.append([x.get("voucher_id", ""), x.get("exam_id", ""), code])

    _append_rows(missing_ws, [["Voucher", "Work / exam", "Missing student code"], *missing_rows])
    _style_header(missing_ws)
    _autosize(missing_ws)

    return wb


@router.get("/teacher_upload_log.xlsx")
@router.get("/stats/teacher_upload_log.xlsx")
def teacher_upload_log_xlsx(
    voucher_id: str = Query(ALL_VOUCHERS),
    _session: dict = Depends(require_teacher),
):
    teacher_id = _teacher_id_from_session(_session)
    payload = _build_teacher_upload_dashboard(teacher_id)

    allowed_vouchers = {str(v.get("voucher_id") or "") for v in payload.get("vouchers", [])}
    selected_voucher = (voucher_id or ALL_VOUCHERS).strip()

    if selected_voucher not in ("", ALL_VOUCHERS) and selected_voucher not in allowed_vouchers:
        raise HTTPException(status_code=404, detail="Voucher/course not found for this teacher.")

    wb = _make_upload_log_workbook(payload, selected_voucher)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    voucher_part = "all_courses" if selected_voucher in ("", ALL_VOUCHERS) else _safe_filename_part(selected_voucher)
    filename = f"student_upload_log_{_safe_filename_part(teacher_id)}_{voucher_part}.xlsx"

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={quote(filename)}"},
    )