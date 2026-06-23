# routes/student_log_routes.py
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.config import RUNS_ROOT
from routes.auth import is_teacher_code

SUBMISSIONS_XLSX = RUNS_ROOT / "student_submissions.xlsx"
SUBMISSIONS_SHEET = "submissions"


def log_student_submission(
    *,
    code: str,
    exam_id: str,
    provider: str,
    ip: str,
    user_agent: str,
    gemini_tokens: int = 0,
    session_role: str | None = None,
) -> None:
    """
    Log ONLY real student submissions.

    Teacher test mode should not be written to student_submissions.xlsx.
    This protects both cases:
    - code is the teacher/admin code
    - the active session role is teacher
    """
    code = (code or "").strip()
    role = (session_role or "").strip().lower()

    if role == "teacher":
        return

    if not code or is_teacher_code(code):
        return

    SUBMISSIONS_XLSX.parent.mkdir(parents=True, exist_ok=True)

    if SUBMISSIONS_XLSX.exists():
        wb = load_workbook(SUBMISSIONS_XLSX)
        ws = wb[SUBMISSIONS_SHEET] if SUBMISSIONS_SHEET in wb.sheetnames else wb.create_sheet(SUBMISSIONS_SHEET)

        if ws.max_row == 0:
            ws.append(["timestamp", "code", "exam_id", "provider", "gemini_tokens", "ip", "user_agent"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SUBMISSIONS_SHEET
        ws.append(["timestamp", "code", "exam_id", "provider", "gemini_tokens", "ip", "user_agent"])

    ws.append([
        datetime.now().isoformat(timespec="seconds"),
        code,
        exam_id,
        provider,
        int(gemini_tokens or 0),
        ip,
        user_agent,
    ])

    wb.save(SUBMISSIONS_XLSX)