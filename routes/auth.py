# routes/auth.py
from __future__ import annotations

import json
import re
import time
from collections import deque
from io import BytesIO
from urllib.parse import quote
from threading import Lock
from typing import Deque, Dict, Set

from fastapi import APIRouter, Depends, File, HTTPException, Request, Response, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

from core.config import RUNS_ROOT
from core.security import (
    clear_session_cookie,
    get_session,
    require_session,
    require_teacher,
    set_profile_session_cookie,
)
from services.student_access import (
    authenticate_student_code,
    create_student_codes,
    get_student_code_record,
    list_student_codes_for_teacher,
    update_student_codes_for_teacher,
)
from services.teacher_profiles import get_teacher
from services.student_work_store import (
    get_student_ocr_daily_usage,
    list_student_work_for_dashboard,
    list_teacher_test_work_for_dashboard,
)

from openpyxl import Workbook, load_workbook

router = APIRouter(prefix="/routes", tags=["auth"])

ALLOWLIST_XLSX = RUNS_ROOT / "student_codes_allowlist.xlsx"
ALLOWLIST_SHEET = "allowlist"


_SAFE_RESULT_NAME_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_\-\. ]{0,160}$")


def _safe_result_name(value: str, fallback: str = "unknown") -> str:
    value = (value or "").strip()
    if not value:
        return fallback
    cleaned = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in value)
    cleaned = cleaned.strip("_")
    return cleaned or fallback


def _student_result_root(student_code: str):
    return RUNS_ROOT / "final_results" / _safe_result_name(student_code or "unknown_student")


def _mask_student_code(code: str) -> str:
    code = (code or "").strip()
    if len(code) <= 6:
        return "student"
    return f"{code[:4]}…{code[-4:]}"


def _safe_result_token(value: str, *, label: str) -> str:
    value = (value or "").strip()
    if not value or "/" in value or "\\" in value or ".." in value:
        raise HTTPException(status_code=400, detail=f"Invalid {label}.")
    if not _SAFE_RESULT_NAME_RE.match(value):
        raise HTTPException(status_code=400, detail=f"Invalid {label}.")
    return value


def _list_previous_student_results(student_code: str) -> list[dict]:
    root = _student_result_root(student_code)
    if not root.exists():
        return []

    results: list[dict] = []

    for meta_path in root.glob("*/*.json"):
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        if not isinstance(meta, dict):
            continue

        # If this result was produced from a saved OCR work item, the dashboard
        # will show it inside that OCR work card instead of as a separate card.
        if str(meta.get("source_work_id") or "").strip():
            continue

        final_file = str(meta.get("final_file") or "").strip()
        if not final_file:
            continue

        result_path = meta_path.parent / final_file
        if not result_path.exists() or result_path.suffix.lower() not in {".pdf", ".tex"}:
            continue

        exam_folder = meta_path.parent.name
        saved_at = str(meta.get("saved_at") or "")
        provider = str(meta.get("provider") or "").strip() or "unknown"
        exam_id = str(meta.get("exam_id") or exam_folder)

        results.append(
            {
                "exam_id": exam_id,
                "exam_folder": exam_folder,
                "provider": provider,
                "saved_at": saved_at,
                "filename": final_file,
                "file_type": result_path.suffix.lower().lstrip("."),
                "download_url": (
                    "/routes/student/result"
                    f"?exam_id={quote(exam_folder)}"
                    f"&filename={quote(final_file)}"
                ),
            }
        )

    results.sort(key=lambda x: x.get("saved_at") or "", reverse=True)
    return results[:30]

# --- Rate limiting -----------------------------------------------------------
# Simple per-IP sliding window for the login endpoint. Not a substitute for a
# real rate limiter, but enough to slow online brute force of student codes
_RL_WINDOW_SECONDS = 60
_RL_MAX_ATTEMPTS = 10
_RL_LOCKOUT_SECONDS = 300
_rl_lock = Lock()
_rl_attempts: Dict[str, Deque[float]] = {}
_rl_lockouts: Dict[str, float] = {}


def _client_ip(request: Request) -> str:
    # If you front this with a reverse proxy, populate X-Forwarded-For carefully
    # — accepting it blindly lets clients spoof the rate-limit key.
    return request.client.host if request.client else "unknown"


def _check_rate_limit(ip: str) -> None:
    now = time.time()
    with _rl_lock:
        locked_until = _rl_lockouts.get(ip, 0)
        if locked_until > now:
            raise HTTPException(
                status_code=429,
                detail=f"Too many attempts. Try again in {int(locked_until - now)}s.",
            )
        if locked_until and locked_until <= now:
            _rl_lockouts.pop(ip, None)


def _record_failed_attempt(ip: str) -> None:
    now = time.time()
    cutoff = now - _RL_WINDOW_SECONDS
    with _rl_lock:
        dq = _rl_attempts.setdefault(ip, deque())
        while dq and dq[0] < cutoff:
            dq.popleft()
        dq.append(now)
        if len(dq) >= _RL_MAX_ATTEMPTS:
            _rl_lockouts[ip] = now + _RL_LOCKOUT_SECONDS
            dq.clear()


def _record_successful_attempt(ip: str) -> None:
    with _rl_lock:
        _rl_attempts.pop(ip, None)
        _rl_lockouts.pop(ip, None)


# --- Allowlist ---------------------------------------------------------------

class LoginRequest(BaseModel):
    code: str


# Teacher account and course request models now live in
# routes/teacher_routes.py.


class CreateStudentCodesRequest(BaseModel):
    count: int = 1
    course_label: str = ""
    student_name: str = ""
    student_email: str = ""
    note: str = ""


def _normalize_code(code: str) -> str:
    return (code or "").strip()


def _ensure_allowlist_file_exists() -> None:
    ALLOWLIST_XLSX.parent.mkdir(parents=True, exist_ok=True)
    if ALLOWLIST_XLSX.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = ALLOWLIST_SHEET
    ws.append(["code"])
    wb.save(ALLOWLIST_XLSX)


def _read_allowlist_codes() -> Set[str]:
    _ensure_allowlist_file_exists()
    wb = load_workbook(ALLOWLIST_XLSX)
    ws = wb[ALLOWLIST_SHEET] if ALLOWLIST_SHEET in wb.sheetnames else wb.active

    codes: Set[str] = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        v = row[0]
        if v is None:
            continue
        s = _normalize_code(str(v))
        if s:
            codes.add(s)
    return codes


def is_allowed_student_code(code: str) -> bool:
    code = _normalize_code(code)
    if not code:
        return False
    return code in _read_allowlist_codes()


@router.post("/login")
def login(req: LoginRequest, request: Request, response: Response):
    ip = _client_ip(request)
    _check_rate_limit(ip)

    code = _normalize_code(req.code)
    if not code:
        _record_failed_attempt(ip)
        raise HTTPException(status_code=400, detail="Missing code")

    student_access = authenticate_student_code(code)
    if not student_access:
        _record_failed_attempt(ip)
        raise HTTPException(status_code=401, detail="Invalid student code")

    set_profile_session_cookie(
        response,
        role="student",
        sub=student_access["code"],
        teacher_id=student_access["teacher_id"],
    )
    _record_successful_attempt(ip)
    return {
        "ok": True,
        "role": "student",
        "teacher_id": student_access["teacher_id"],
        "subject": student_access.get("subject", "math"),
        "course_label": student_access.get("course_label", ""),
    }

# Teacher registration, login, profile, password, and course routes now live in:
# routes/teacher_routes.py


def _teacher_id_or_400(session: dict) -> str:
    teacher_id = (session.get("teacher_id") or session.get("sub") or "").strip()
    if not teacher_id or teacher_id == "teacher":
        raise HTTPException(status_code=400, detail="Legacy teacher sessions cannot manage student codes.")
    return teacher_id


def _active_voucher_id_for_teacher(teacher_id: str) -> str:
    teacher = get_teacher(teacher_id) or {}
    active_course = teacher.get("active_course") if isinstance(teacher.get("active_course"), dict) else {}
    return str(
        active_course.get("voucher_id")
        or teacher.get("voucher_id")
        or teacher.get("voucher_code")
        or teacher.get("voucher_hash")
        or ""
    ).strip()


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_excel_header(value) -> str:
    key = _cell_text(value).lower()
    key = re.sub(r"[^a-z0-9]+", "_", key).strip("_")
    aliases = {
        "student": "student_name",
        "student_name": "student_name",
        "name": "student_name",
        "email": "student_email",
        "student_email": "student_email",
        "course": "course_label",
        "course_group": "course_label",
        "course_label": "course_label",
        "group": "course_label",
        "note": "note",
        "notes": "note",
        "status": "status",
        "code": "code",
        "student_code": "code",
    }
    return aliases.get(key, key)


def _student_codes_workbook(codes: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "student_codes"

    headers = [
        "code",
        "student_name",
        "student_email",
        "course_label",
        "note",
        "status",
        "uses",
        "last_used_at",
        "created_at",
    ]
    ws.append(headers)

    for rec in codes:
        ws.append([rec.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 34)

    return wb


def _rows_from_uploaded_student_codes_xlsx(content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_excel_header(x) for x in rows[0]]
    out: list[dict[str, str]] = []
    for raw in rows[1:]:
        row = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            row[header] = _cell_text(raw[i] if i < len(raw) else None)
        if any(row.values()):
            out.append(row)
    return out


@router.get("/teacher/student_codes")
def teacher_student_codes(_session: dict = Depends(require_teacher)):
    teacher_id = _teacher_id_or_400(_session)
    voucher_id = _active_voucher_id_for_teacher(teacher_id)
    return {
        "ok": True,
        "teacher_id": teacher_id,
        "voucher_id": voucher_id,
        "codes": list_student_codes_for_teacher(teacher_id, voucher_id=voucher_id),
    }


@router.post("/teacher/student_codes")
def teacher_create_student_codes(req: CreateStudentCodesRequest, _session: dict = Depends(require_teacher)):
    teacher_id = _teacher_id_or_400(_session)

    codes = create_student_codes(
        teacher_id=teacher_id,
        count=req.count,
        course_label=req.course_label,
        note=req.note,
        student_name=req.student_name,
        student_email=req.student_email,
    )
    return {"ok": True, "teacher_id": teacher_id, "codes": codes}


@router.get("/teacher/student_codes.xlsx")
def teacher_student_codes_xlsx(_session: dict = Depends(require_teacher)):
    teacher_id = _teacher_id_or_400(_session)
    voucher_id = _active_voucher_id_for_teacher(teacher_id)
    codes = list_student_codes_for_teacher(teacher_id, voucher_id=voucher_id)
    wb = _student_codes_workbook(codes)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"student_codes_{teacher_id}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={quote(filename)}"},
    )


@router.post("/teacher/student_codes/import")
async def teacher_import_student_codes(
    file: UploadFile = File(...),
    _session: dict = Depends(require_teacher),
):
    teacher_id = _teacher_id_or_400(_session)
    name = (file.filename or "").lower()
    if not name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx file exported from the student-codes table.")

    content = await file.read()
    rows = _rows_from_uploaded_student_codes_xlsx(content)
    result = update_student_codes_for_teacher(teacher_id, rows)
    return {"ok": True, "teacher_id": teacher_id, **result}


@router.get("/student/dashboard")
def student_dashboard(_session: dict = Depends(require_session)):
    role = _session.get("role")

    if role == "teacher":
        teacher_id = str(
            _session.get("teacher_id")
            or _session.get("sub")
            or ""
        ).strip()

        if not teacher_id:
            raise HTTPException(
                status_code=403,
                detail=(
                    "Missing teacher profile "
                    "in session."
                ),
            )

        teacher = (
            get_teacher(teacher_id)
            or {}
        )

        active_course = (
            teacher.get("active_course")
            if isinstance(
                teacher.get("active_course"),
                dict,
            )
            else {}
        )

        active_voucher_id = str(
            active_course.get("voucher_id")
            or active_course.get("course_id")
            or teacher.get("voucher_id")
            or teacher.get("active_course_id")
            or ""
        ).strip()

        teacher_test_results = (
            list_teacher_test_work_for_dashboard(
                teacher_id,
                voucher_id=active_voucher_id,
            )
            if active_voucher_id
            else []
        )

        active_course_label = str(
            active_course.get("course_label")
            or teacher.get("course_label")
            or ""
        ).strip()

        active_subject = str(
            active_course.get("subject")
            or teacher.get("subject")
            or "math"
        ).strip()

        active_subject_label = str(
            active_course.get(
                "subject_label"
            )
            or teacher.get(
                "subject_label"
            )
            or active_subject
            or "Math"
        ).strip()

        return {
            "ok": True,
            "role": "teacher",
            "is_teacher_test": True,
            "teacher": {
                "teacher_id": teacher_id,
                "name": (
                    teacher.get("name")
                    or teacher.get("username")
                    or "Teacher"
                ),
                "email": (
                    teacher.get("email")
                    or ""
                ),
                "subject": active_subject,
                "subject_label": (
                    active_subject_label
                ),
            },
            "course": {
                "course_label": (
                    "Teacher test mode"
                ),
                "subject": active_subject,
                "subject_label": (
                    active_subject_label
                ),
                "note": (
                    f"Active course: "
                    f"{active_course_label}"
                    if active_course_label
                    else ""
                ),
                "voucher_id": (
                    active_voucher_id
                ),
            },
            "student": None,
            "previous_results": (
                teacher_test_results[:40]
            ),
        }

    if role != "student":
        raise HTTPException(status_code=403, detail="Student dashboard requires a student session.")

    student_code = str(_session.get("sub") or "").strip()
    if not student_code:
        raise HTTPException(status_code=403, detail="Missing student code in session.")

    record = get_student_code_record(student_code)
    if not record:
        raise HTTPException(status_code=404, detail="Student code record not found.")

    teacher_id = str(
        record.get("teacher_id")
        or ""
    ).strip()

    teacher = (
        get_teacher(teacher_id)
        or {}
    )

    voucher_id = str(
        record.get("voucher_id")
        or record.get("voucher_hash")
        or record.get("voucher_code")
        or ""
    ).strip()

    ocr_quota = (
        get_student_ocr_daily_usage(
            student_code,
            teacher_id=teacher_id,
            voucher_id=voucher_id,
        )
    )

    return {
        "ok": True,
        "role": "student",
        "is_teacher_test": False,
        "student": {
            "student_code_display": _mask_student_code(student_code),
            "uses": int(record.get("uses") or 0),
            "last_used_at": record.get("last_used_at") or "",
        },
        "teacher": {
            "teacher_id": teacher_id,
            "name": record.get("teacher_name") or teacher.get("name") or "",
            "email": record.get("teacher_email") or teacher.get("email") or "",
        },
        "course": {
            "course_label": record.get("course_label") or "Course",
            "subject": record.get("subject") or teacher.get("subject") or "math",
            "subject_label": record.get("subject_label") or teacher.get("subject_label") or record.get("subject") or "Math",
            "note": record.get("note") or "",
        },
        "ocr_quota": ocr_quota,

        # Canonical student-visible history now lives in:
        # data/teachers/<teacher>/courses/<voucher>/
        # students/<student>/<work_id>/
        "previous_results": (
            list_student_work_for_dashboard(
                student_code
            )[:40]
        ),
    }


@router.get("/student/result")
def student_result_download(
    exam_id: str,
    filename: str,
    _session: dict = Depends(require_session),
):
    if _session.get("role") != "student":
        raise HTTPException(status_code=403, detail="Only students can download their saved feedback here.")

    student_code = str(_session.get("sub") or "").strip()
    if not student_code:
        raise HTTPException(status_code=403, detail="Missing student code in session.")

    exam_folder = _safe_result_token(exam_id, label="exam_id")
    safe_filename = _safe_result_token(filename, label="filename")

    if not safe_filename.lower().endswith((".pdf", ".tex")):
        raise HTTPException(status_code=400, detail="Unsupported result file type.")

    root = _student_result_root(student_code).resolve()
    path = (root / exam_folder / safe_filename).resolve()

    try:
        path.relative_to(root)
    except ValueError:
        raise HTTPException(status_code=403, detail="Invalid result path.")

    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Result file not found.")

    if path.suffix.lower() == ".pdf":
        media_type = "application/pdf"
    else:
        media_type = "text/plain; charset=utf-8"

    return FileResponse(
        path=str(path),
        media_type=media_type,
        filename=path.name,
    )


@router.post("/logout")
def logout(response: Response):
    clear_session_cookie(response)
    return {"ok": True}


@router.get("/me")
def me(request: Request):
    s = get_session(request)
    if not s:
        return {"ok": False, "role": None}
    # Never leak the student code back to the client; the cookie itself is
    # the source of truth.
    teacher_id = s.get("teacher_id") or (s.get("sub") if s.get("role") == "teacher" else "")
    return {"ok": True, "role": s.get("role"), "teacher_id": teacher_id}
