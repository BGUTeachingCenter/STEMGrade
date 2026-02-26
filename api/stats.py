# api/stats.py
from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Header
from openpyxl import load_workbook

from core.config import RUNS_ROOT
from core.security import require_teacher_password

router = APIRouter(prefix="/api", tags=["stats"])

SUBMISSIONS_XLSX = RUNS_ROOT / "student_submissions.xlsx"
SHEET = "submissions"


def _read_rows() -> list[dict[str, Any]]:
    if not SUBMISSIONS_XLSX.exists():
        return []

    wb = load_workbook(SUBMISSIONS_XLSX)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    out = []
    for r in rows[1:]:
        d = {}
        for i, key in enumerate(header):
            if not key:
                continue
            d[key] = r[i] if i < len(r) else None
        out.append(d)
    return out


@router.get("/teacher_dashboard")
def teacher_dashboard(x_teacher_password: str | None = Header(None)):
    require_teacher_password(x_teacher_password)

    rows = _read_rows()

    # 1) usage per day (submissions/day)
    per_day = Counter()
    for r in rows:
        ts = r.get("timestamp")
        if not ts:
            continue
        # timestamp is ISO string like 2026-02-26T14:33:36
        day = str(ts)[:10]
        per_day[day] += 1
    usage_per_day = [{"day": d, "count": per_day[d]} for d in sorted(per_day.keys())]

    # 2) how many codes submitted each exam_id (unique codes + total submissions)
    exam_codes = defaultdict(set)
    exam_subs = Counter()
    for r in rows:
        exam_id = (r.get("exam_id") or "").strip()
        code = (r.get("code") or "").strip()
        if not exam_id:
            continue
        exam_subs[exam_id] += 1
        if code:
            exam_codes[exam_id].add(code)

    codes_per_exam = []
    for exam_id in sorted(exam_subs.keys()):
        codes_per_exam.append({
            "exam_id": exam_id,
            "unique_codes": len(exam_codes[exam_id]),
            "submissions": int(exam_subs[exam_id]),
        })

    # 3) gemini tokens per code (aggregate)
    tokens_by_code = Counter()
    for r in rows:
        provider = (r.get("provider") or "").strip().lower()
        if provider not in ("google", "gemini", "google_ai_studio", "aistudio"):
            continue
        code = (r.get("code") or "").strip()
        if not code:
            continue
        tok = r.get("gemini_tokens")
        try:
            tok_i = int(tok or 0)
        except Exception:
            tok_i = 0
        tokens_by_code[code] += tok_i

    gemini_tokens_per_code = [{"code": c, "tokens": int(tokens_by_code[c])} for c in sorted(tokens_by_code.keys())]

    return {
        "usage_per_day": usage_per_day,
        "codes_per_exam": codes_per_exam,
        "gemini_tokens_per_code": gemini_tokens_per_code,
    }