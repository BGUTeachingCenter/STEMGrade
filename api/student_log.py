# api/student_log.py
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.config import RUNS_ROOT
from api.auth import is_teacher_code

SUBMISSIONS_XLSX = RUNS_ROOT / "student_submissions.xlsx"
SUBMISSIONS_SHEET = "submissions"


def log_student_submission(
    *,
    code: str,
    exam_id: str,
    provider: str,
    ip: str,
    user_agent: str,
    gemini_tokens: int = 0,
) -> None:
    """
    Log ONLY student submissions. Admin codes are omitted by design.
    """
    code = (code or "").strip()
    if not code or is_teacher_code(code):
        return

    SUBMISSIONS_XLSX.parent.mkdir(parents=True, exist_ok=True)

    if SUBMISSIONS_XLSX.exists():
        wb = load_workbook(SUBMISSIONS_XLSX)
        ws = wb[SUBMISSIONS_SHEET] if SUBMISSIONS_SHEET in wb.sheetnames else wb.create_sheet(SUBMISSIONS_SHEET)
        # If old file exists without the new column, we still append safely.
        if ws.max_row == 0:
            ws.append(["timestamp", "code", "exam_id", "provider", "gemini_tokens", "ip", "user_agent"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SUBMISSIONS_SHEET
        ws.append(["timestamp", "code", "exam_id", "provider", "gemini_tokens", "ip", "user_agent"])

    ws.append([
        datetime.now().isoformat(timespec="seconds"),
        code,
        exam_id,
        provider,
        int(gemini_tokens or 0),
        ip,
        user_agent,
    ])
    wb.save(SUBMISSIONS_XLSX)