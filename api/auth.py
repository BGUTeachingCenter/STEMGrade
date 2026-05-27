# api/auth.py
from __future__ import annotations

import time
from collections import deque
from threading import Lock
from typing import Deque, Dict, Set

from fastapi import APIRouter, HTTPException, Request, Response
from pydantic import BaseModel

from core.config import RUNS_ROOT
from core.security import (
    clear_session_cookie,
    get_session,
    set_session_cookie,
    verify_teacher_password,
)

from openpyxl import Workbook, load_workbook

router = APIRouter(prefix="/api", tags=["auth"])

ALLOWLIST_XLSX = RUNS_ROOT / "student_codes_allowlist.xlsx"
ALLOWLIST_SHEET = "allowlist"

# --- Rate limiting -----------------------------------------------------------
# Simple per-IP sliding window for the login endpoint. Not a substitute for a
# real rate limiter, but enough to slow online brute force of student codes
# and the teacher password.
_RL_WINDOW_SECONDS = 60
_RL_MAX_ATTEMPTS = 10
_RL_LOCKOUT_SECONDS = 300
_rl_lock = Lock()
_rl_attempts: Dict[str, Deque[float]] = {}
_rl_lockouts: Dict[str, float] = {}


def _client_ip(request: Request) -> str:
    # If you front this with a reverse proxy, populate X-Forwarded-For carefully
    # — accepting it blindly lets clients spoof the rate-limit key.
    return request.client.host if request.client else "unknown"


def _check_rate_limit(ip: str) -> None:
    now = time.time()
    with _rl_lock:
        locked_until = _rl_lockouts.get(ip, 0)
        if locked_until > now:
            raise HTTPException(
                status_code=429,
                detail=f"Too many attempts. Try again in {int(locked_until - now)}s.",
            )
        if locked_until and locked_until <= now:
            _rl_lockouts.pop(ip, None)


def _record_failed_attempt(ip: str) -> None:
    now = time.time()
    cutoff = now - _RL_WINDOW_SECONDS
    with _rl_lock:
        dq = _rl_attempts.setdefault(ip, deque())
        while dq and dq[0] < cutoff:
            dq.popleft()
        dq.append(now)
        if len(dq) >= _RL_MAX_ATTEMPTS:
            _rl_lockouts[ip] = now + _RL_LOCKOUT_SECONDS
            dq.clear()


def _record_successful_attempt(ip: str) -> None:
    with _rl_lock:
        _rl_attempts.pop(ip, None)
        _rl_lockouts.pop(ip, None)


# --- Allowlist ---------------------------------------------------------------

class LoginRequest(BaseModel):
    code: str


def _normalize_code(code: str) -> str:
    return (code or "").strip()


def _ensure_allowlist_file_exists() -> None:
    ALLOWLIST_XLSX.parent.mkdir(parents=True, exist_ok=True)
    if ALLOWLIST_XLSX.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = ALLOWLIST_SHEET
    ws.append(["code"])
    wb.save(ALLOWLIST_XLSX)


def _read_allowlist_codes() -> Set[str]:
    _ensure_allowlist_file_exists()
    wb = load_workbook(ALLOWLIST_XLSX)
    ws = wb[ALLOWLIST_SHEET] if ALLOWLIST_SHEET in wb.sheetnames else wb.active

    codes: Set[str] = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        v = row[0]
        if v is None:
            continue
        s = _normalize_code(str(v))
        if s:
            codes.add(s)
    return codes


def is_teacher_code(code: str) -> bool:
    return verify_teacher_password(_normalize_code(code))


def is_allowed_student_code(code: str) -> bool:
    code = _normalize_code(code)
    if not code:
        return False
    if is_teacher_code(code):
        return False
    return code in _read_allowlist_codes()


@router.post("/login")
def login(req: LoginRequest, request: Request, response: Response):
    ip = _client_ip(request)
    _check_rate_limit(ip)

    code = _normalize_code(req.code)
    if not code:
        _record_failed_attempt(ip)
        raise HTTPException(status_code=400, detail="Missing code")

    if is_teacher_code(code):
        set_session_cookie(response, role="teacher", sub="teacher")
        _record_successful_attempt(ip)
        return {"ok": True, "role": "teacher"}

    if not is_allowed_student_code(code):
        _record_failed_attempt(ip)
        raise HTTPException(status_code=401, detail="Invalid code")

    set_session_cookie(response, role="student", sub=code)
    _record_successful_attempt(ip)
    return {"ok": True, "role": "student"}


@router.post("/logout")
def logout(response: Response):
    clear_session_cookie(response)
    return {"ok": True}


@router.get("/me")
def me(request: Request):
    s = get_session(request)
    if not s:
        return {"ok": False, "role": None}
    # Never leak the student code back to the client; the cookie itself is
    # the source of truth.
    return {"ok": True, "role": s.get("role")}
